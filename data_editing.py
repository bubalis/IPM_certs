#!/usr/bin/env python
# coding: utf-8
'''This script assembles the workbook notes into a single csv file.'''


import openpyxl as xl
import pandas as pd
import os


wb_path='Cert_coding.xlsx'
wb=xl.load_workbook(wb_path)






def all_unique_codes(wb):
    unique_names=[]
    for sheet in wb.sheetnames:
        if sheet in ['Master', "Excluded"]:
            pass
        else:
            df=pd.read_excel(wb_path, sheet_name=sheet)
            df=df.fillna('')
            for column in df.columns[8:]:
                li=list(df[column].unique())
                #print(li)
                unique_names+=li
                
    return list(set([str(n).lower() for n in unique_names if n]))




names=all_unique_codes(wb)








def workbook_editor(wb, corrections):
    for sheet in wb.sheetnames:
            if sheet in ['Master', 'Excluded']:
                pass
            else:
                df=pd.read_excel(wb_path, sheet_name=sheet)
                df=df.fillna('')
                for column in df.columns[8:]:
                    li=list(df[column].unique())
                    #print(li)
                    unique_names+=li
                
    return list(set([str(n).lower() for n in unique_names if n]))






corrections={'on-site research': 'on-farm research',
 'training for workers who handle pesticides': 'workers who handle pesticides are: trained',
 'medical testing for workers who handle dangerous chemicals': 'workers who handle pesticides receive: medical testing',
 'annual medical examination for workers who handle pesticides': 'workers who handle pesticides receive: annual medical evaluations',
 'workers trained to use pesiticides': 'workers who handle pesticides are: trained',
 'adequate': 'proper',
'records of weather during spraying': 'weather records',
'records kept of scouting': 'monitoring records',

'records of all pesticide applications': 'records of all pesticide use',
 'written records for monitoring': 'monitoring records',
'monitoring records are kept': 'monitoring records',
'detailed pesticide application records': 'records of all pesticide use',
'record-keeping of all pesticide applications': 'records of all pesticide use',
'records of all pesticides used': 'records of all pesticide use',
'monitoring records are kept': 'records of all pesticide use',
'records are used to modify plan': 'records used in planning', 
'uses records to devise strategies': 'records used in planning', 
'records of pest/disease monitoring': 'monitoring records',
'scouting and weather records': 'monitoring records, weather records',
'uses records to devise strategies, records of efficacy, weather, pest monitoring, thresholds':
'efficacy records, weather records, monitoring records, records used in planning',
'use pesticides': 'handle pesticides',
'workers trained to use ppe, handle pesticides': 'workers trained to use ppe, workers trained to handle pesticides',
'workers trained in pesticide safety': 'workers trained to handle pesticides',
'all workers trained in pesticide safety': 'workers trained to handle pesticides',
'trained to use ppe': 'workers trained to use ppe',
'employees': 'workers',
'training for workers who handle pesticides': 'workers trained to handle pesticides',
'farmers trained in pesticide application': 'workers trained to handle pesticides',
'workers who handle pesticides are trained': 'workers trained to handle pesticides',
 'workers who handle pesticides are: trained': 'workers trained to handle pesticides',
'employees trained in pesticides': 'workers trained to handle pesticides', 
'workers who handle pesticides: are trained; have access to ppe': 'workers trained to handle pesticides, workers have proper ppe',
'threshold ': 'thresholds',
'weeds scouted 1x/season': 'scouting weeds 1x/season',
'monitor weather' : 'weather monitoring',
'monitor invasive plants': 'invasive plant monitoring',
'monitor crop at least 2x': 'crop monitoring, at least 2x/season',
'monitor insect numbers (specific insect)': 'pest monitoring (specific insect)',
'monitor high risks for resistance': 'monitoring for resistance of pests',
'helath': 'health', 
're-entry periods': 're-entry intervals after pesticide application',
 're-entry periods observed':  're-entry intervals after pesticide application',
're-entry periods': 're-entry intervals after pesticide application',
'roation': 'rotation',
'avgue': 'vague',
'moa resistance': 'moa rotation',

                 
}




def check_replace(string, mapping):
    for k, v in mapping.items():
        string=str(string).replace(k, v)
    return string




# In[90]:


def corrector(string):
    return check_replace(string, corrections)





df=pd.DataFrame()
for sheet in wb.sheetnames:
    if sheet=='Master':
        pass
    else:
        df2=pd.read_excel(wb_path, sheet_name=sheet)
        df2=df2.dropna(how='all')
        df2=df2.fillna('')
        for col in df2.columns[8:]:
            df2[col]=df2[col].apply(lambda x: str(x).lower())
            df2[col]=df2[col].apply(corrector)
            df2['CertName']=sheet
        df=df.append(df2)

df.rename(columns={'Required Practice': 'Pesticide Practices'}, inplace=True)



for col in df.columns[5:]:
    df[col]=df[col].apply(corrector)




df.to_csv('all_data.csv')


